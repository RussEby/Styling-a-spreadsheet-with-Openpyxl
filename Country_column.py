# import the needed items
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill


def main(old_file, new_file):
    # load the plain file
    wb = load_workbook(filename=old_file)

    # select the active page
    ws = wb.active

    # the countries are in the first column
    country_column = "A"

    # to make things a little clearer these are the colors I'm using for styling
    babyBlue = "00CCFFFF"

    # Here we are defining the border style
    thick_border = Side(border_style="thick")

    # iterate over the Country column skipping the header row
    for cell in ws[country_column][1:]:
        # Formating each cell
        cell.fill = PatternFill(
            start_color=babyBlue, end_color=babyBlue, fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="right")
        cell.border = Border(right=thick_border)

    # save the OpenPyXL Workbook
    wb.save(new_file)


if __name__ == "__main__":
    main(old_file="data/header.xlsx", new_file="data/country_column.xlsx")
