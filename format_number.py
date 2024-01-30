# import the needed items
from openpyxl import load_workbook


def main(old_file, new_file):
    # load the plain file
    wb = load_workbook(filename=old_file)

    # select the active page
    ws = wb.active

    # 2 types of numeric columns: no decimals and 2 decimals
    number_col_0_dec = ["C", "D", "I"]

    # iterate over the number columns
    for col in ws["C:T"]:
        # iterate over the cells in that column, skip the header row
        for cell in col[1:]:
            # if the cell has value and is type str
            if cell.value and isinstance(cell.value, str):
                # replace the comma with a period
                cell.value = float(cell.value.replace(",", "."))

            # apply format
            cell.number_format = "#,##0" if col in number_col_0_dec else "#,##0.00"

    # save the OpenPyXL Workbook
    wb.save(new_file)


if __name__ == "__main__":
    main(old_file="data/country_column.xlsx", new_file="data/format_number.xlsx")
