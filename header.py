# import the needed items
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def autofit(sheet):
    # Iterate over the solumns
    for column_cells in sheet.columns:
        # determine the longest item in each column
        max_length = max(len(cell.value or "") for cell in column_cells)

        # I'm adding an extra 2 since I like it a bit roomy
        max_length = (max_length + 2) * 1.2

        # set the width of the column to the max_length
        sheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = max_length


def format_header_cell(cell):
    # to make things a little clearer these are the colors I'm using for styling
    blue = "000000FF"
    lightBlue = "0099CCFF"
    black = "00000000"
    white = "00FFFFFF"

    # Here we are defining the border style
    thin_border = Side(border_style="thin", color=black)
    double_border = Side(border_style="double", color=black)

    # apply a fill to the cell
    cell.fill = PatternFill(start_color=blue, end_color=lightBlue, fill_type="solid")

    # apply the font styles
    cell.font = Font(name="Tahoma", size=12, color=white, bold=True)

    # apply an alignment
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # apply the border
    cell.border = Border(
        top=double_border, left=thin_border, right=thin_border, bottom=double_border
    )


def main(old_file, new_file):
    # load the plain file
    wb = load_workbook(filename=old_file)

    # select the active page
    ws = wb.active

    # the column header is in the first row
    column_header_row = 1

    # select the column headers
    for cell in ws[column_header_row]:
        format_header_cell(cell)

    # Use the user defined function to set the column width
    autofit(ws)

    # Adjust Column 'H' due to it's width
    ws.column_dimensions["H"].width = ws.column_dimensions["H"].width * 0.6

    # save the OpenPyXL Workbook
    wb.save(new_file)


if __name__ == "__main__":
    main(old_file="data/cleaned.xlsx", new_file="data/header.xlsx")
