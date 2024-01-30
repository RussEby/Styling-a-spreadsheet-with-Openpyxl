# import the needed items
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side


def format_header_cell(cell):
    # to make things a little clearer these are the colors I'm using for styling
    blue = "000000FF"
    lightBlue = "0099CCFF"
    black = "00000000"
    white = "00FFFFFF"

    # Here we are defining the border style
    thin_border = Side(border_style="thin", color=blue)
    double_border = Side(border_style="double", color=lightBlue)

    cell.fill = PatternFill(start_color=black, end_color=black, fill_type="solid")
    cell.font = Font(name="Tahoma", size=12, color=white, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        top=double_border, left=thin_border, right=thin_border, bottom=double_border
    )


def main(old_file, new_file):
    # load the plain file
    wb = load_workbook(filename=old_file)

    # add a summary sheet
    rankings_sheet = wb.create_sheet(title="Rankings")

    # Header Row
    rankings_sheet["A1"] = "Ranking"
    format_header_cell(rankings_sheet["A1"])
    rankings_sheet.column_dimensions["A"].width = 25

    format_header_cell(rankings_sheet["B1"])
    rankings_sheet["B1"] = "Country"
    rankings_sheet.column_dimensions["B"].width = 20

    format_header_cell(rankings_sheet["C1"])
    rankings_sheet["C1"] = "Values"
    rankings_sheet.column_dimensions["C"].width = 15

    # Largest Population
    rankings_sheet["A2"] = "Largest population"
    rankings_sheet[
        "B2"
    ] = "=INDEX('Country Data'!A2:A229, MATCH(MAX('Country Data'!C2:C228), 'Country Data'!C2:C228, 0))"
    rankings_sheet["C2"] = "=MAX('Country Data'!C2:C228)"
    rankings_sheet["C2"].number_format = "#,###"

    # Largest Population
    rankings_sheet["A3"] = "Smallest population"
    rankings_sheet[
        "B3"
    ] = "=INDEX('Country Data'!A2:A229, MATCH(MIN('Country Data'!C2:C228), 'Country Data'!C2:C228, 0))"
    rankings_sheet["C3"] = "=MIN('Country Data'!C2:C228)"
    rankings_sheet["C3"].number_format = "#,###"

    # Largest Area
    rankings_sheet["A4"] = "Largest area"
    rankings_sheet[
        "B4"
    ] = "=INDEX('Country Data'!A2:A229, MATCH(MAX('Country Data'!D2:D228), 'Country Data'!D2:D228, 0))"
    rankings_sheet["C4"] = "=MAX('Country Data'!D2:D228)"
    rankings_sheet["C4"].number_format = "#,###"

    # Smallest Area
    rankings_sheet["A5"] = "Smallest area"
    rankings_sheet["B5"] = "=INDEX('Country Data'!A2:A229, MATCH(MIN('Country Data'!D2:D228), 'Country Data'!D2:D228, 0))"
    rankings_sheet["C5"] = "=MIN('Country Data'!D2:D228)"
    rankings_sheet["C5"].number_format = "#,###"

    # save the OpenPyXL Workbook
    wb.save(new_file)


if __name__ == "__main__":
    main(old_file="data/format_number.xlsx", new_file="data/rankings_tab.xlsx")
