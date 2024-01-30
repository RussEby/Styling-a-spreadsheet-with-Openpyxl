# import the needed items
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

def format_header_cell(cell):
    # to make things a little clearer these are the colors I'm using for styling
    blue = "000000FF"
    lightBlue = "0099CCFF"
    black = "00000000"
    white = "00FFFFFF"

    # Here we are defining the border style
    thin_border = Side(border_style="thin", color=black)
    double_border = Side(border_style="double", color=black)

    # apply a fill to the cell
    cell.fill = PatternFill(start_color=blue, end_color=lightBlue, fill_type="solid")

    # apply the font styles
    cell.font = Font(name="Tahoma", size=12, color=white, bold=True)

    # apply an alignment
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # apply the border
    cell.border = Border(
        top=double_border, left=thin_border, right=thin_border, bottom=double_border
    )
    

def main(old_file, new_file):
    # load the plain file
    workbook = load_workbook(filename=old_file)

    # add a summary sheet
    summary_sheet = workbook.create_sheet(title="Summary")

    # Region header
    summary_sheet['A1'].value = 'Region'
    format_header_cell(summary_sheet['A1'])

    # per OpenPyXL's documention, any formula that wasn't in the initial specification
    # must be prefixed with `_xlfn.`
    summary_sheet['A2'] = ArrayFormula(
        "A2:A12", 
        "=_xlfn.UNIQUE(_xlfn.SORT('Country Data'!B2:B228))"
        )

    summary_sheet.column_dimensions['A'].width = 22

    # Count of Country
    summary_sheet['B1'].value = '# of Countries'
    format_header_cell(summary_sheet['B1'])

    for col in summary_sheet['B2:B12']:
        for cell in col:
            cell.value = f"=COUNTIF('Country Data'!$B$2:$B$228, A{cell.row})"
            cell.number_format = "#,###"

    summary_sheet.column_dimensions['B'].width = 15

    # Population
    summary_sheet['C1'].value = 'Population'
    format_header_cell(summary_sheet['C1'])

    for col in summary_sheet['C2:C12']:
        for cell in col:
            cell.value = f"=SUMIF('Country Data'!$B$2:$B$228, A{cell.row}, 'Country Data'!$C$2:$C$228)"
            cell.number_format = "#,###"

    summary_sheet.column_dimensions['C'].width = 15

    # Area
    summary_sheet['D1'].value = 'Area'
    format_header_cell(summary_sheet['D1'])

    for col in summary_sheet['D2:D12']:
        for cell in col:
            cell.value = f"=SUMIF('Country Data'!$B$2:$B$228, A{cell.row}, 'Country Data'!$D$2:$CD$228)"
            cell.number_format = "#,###"
    
    summary_sheet.column_dimensions['D'].width = 15

    # save the OpenPyXL Workbook
    workbook.save(new_file)


if __name__ == "__main__":
    main(old_file="data/rankings_tab.xlsx", new_file="data/summary_tab.xlsx")